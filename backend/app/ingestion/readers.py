import csv
import re
from dataclasses import dataclass
from io import BytesIO, StringIO


HEADER_KEYWORDS = {
    "date",
    "txn date",
    "transaction date",
    "value date",
    "posted date",
    "narration",
    "description",
    "details",
    "merchant",
    "amount",
    "debit",
    "credit",
    "withdrawal",
    "deposit",
    "balance",
    "type",
}


@dataclass
class ParsedSheet:
    sheet_name: str | None
    header_row_index: int
    headers: list[str]
    rows: list[dict]
    score: int


def _decode_text(content: bytes) -> str:
    for encoding in ("utf-8-sig", "utf-8", "latin-1"):
        try:
            return content.decode(encoding)
        except UnicodeDecodeError:
            continue
    return content.decode("utf-8", errors="ignore")


def _normalize_header(cell: object) -> str:
    value = "" if cell is None else str(cell).strip()
    value = value.lower().replace("_", " ").replace("-", " ")
    value = re.sub(r"[\(\)\[\]\.`,:;/]+", " ", value)
    return " ".join(value.split())


def _score_header_row(row: list[object]) -> int:
    normalized = [_normalize_header(cell) for cell in row if _normalize_header(cell)]
    if not normalized:
        return 0
    keyword_score = sum(2 for cell in normalized if cell in HEADER_KEYWORDS)
    diversity_score = min(len(set(normalized)), 6)
    return keyword_score + diversity_score


def _build_rows(headers: list[str], data_rows: list[list[object]]) -> list[dict]:
    rows: list[dict] = []
    for row in data_rows:
        if not any(cell not in (None, "") for cell in row):
            continue
        padded = list(row) + [None] * max(0, len(headers) - len(row))
        rows.append({headers[index]: padded[index] for index in range(len(headers)) if headers[index]})
    return rows


def _detect_best_sheet(sheet_name: str | None, matrix: list[list[object]]) -> ParsedSheet:
    best_index = 0
    best_score = -1
    scan_limit = min(len(matrix), 20)
    for row_index in range(scan_limit):
        score = _score_header_row(matrix[row_index])
        if score > best_score:
            best_score = score
            best_index = row_index

    headers = [_normalize_header(cell) for cell in matrix[best_index]]
    data_rows = matrix[best_index + 1 :]
    rows = _build_rows(headers, data_rows)
    return ParsedSheet(
        sheet_name=sheet_name,
        header_row_index=best_index,
        headers=headers,
        rows=rows,
        score=best_score,
    )


def read_csv_rows(content: bytes) -> ParsedSheet:
    text = _decode_text(content)
    reader = csv.reader(StringIO(text))
    matrix = [row for row in reader]
    return _detect_best_sheet(None, matrix)


def read_xlsx_rows(content: bytes) -> ParsedSheet:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("XLSX support requires openpyxl to be installed.") from exc

    workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
    best_sheet: ParsedSheet | None = None
    for sheet_name in workbook.sheetnames:
        worksheet = workbook[sheet_name]
        matrix = [list(row) for row in worksheet.iter_rows(values_only=True)]
        if not matrix:
            continue
        candidate = _detect_best_sheet(sheet_name, matrix)
        if best_sheet is None or candidate.score > best_sheet.score or len(candidate.rows) > len(best_sheet.rows):
            best_sheet = candidate

    if best_sheet is None:
        return ParsedSheet(sheet_name=None, header_row_index=0, headers=[], rows=[], score=0)
    return best_sheet


def read_xls_rows(content: bytes) -> ParsedSheet:
    try:
        import xlrd
    except ImportError as exc:
        raise RuntimeError("XLS support requires xlrd to be installed.") from exc

    workbook = xlrd.open_workbook(file_contents=content)
    best_sheet: ParsedSheet | None = None
    for sheet in workbook.sheets():
        matrix = [[sheet.cell_value(row, col) for col in range(sheet.ncols)] for row in range(sheet.nrows)]
        if not matrix:
            continue
        candidate = _detect_best_sheet(sheet.name, matrix)
        if best_sheet is None or candidate.score > best_sheet.score or len(candidate.rows) > len(best_sheet.rows):
            best_sheet = candidate

    if best_sheet is None:
        return ParsedSheet(sheet_name=None, header_row_index=0, headers=[], rows=[], score=0)
    return best_sheet
